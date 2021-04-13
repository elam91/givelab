import openpyxl
import time
import schedule
from airtable import Airtable
from selenium import webdriver
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.firefox.options import Options

def excel_to_dict(excel_path, headers=[]):
    wb = openpyxl.load_workbook(excel_path)
    sheet = wb['Sheet1']
    result_dict = []
    for row in range(2, sheet.max_row +1):
        line = dict()
        for header in headers:
            cell_value = sheet.cell(column=(int(headers.index(header)+1)),row=int(row)).value
            if type(cell_value) is int:
                cell_value = str(cell_value)
            elif cell_value is None:
                cell_value = ''
            line[header] = cell_value
        result_dict.append(line)
    return result_dict


def go_to_givelab(username,user_password,giveaway_url):
    options = Options()
    options.headless = True
    browser = webdriver.Firefox(options=options)
    browser.get("https://givelab.com/login?ref=https%3A%2F%2Fgivelab.com%2F")

    try:
        element_present = EC.presence_of_element_located((By.ID, 'loginEmail'))
        email = WebDriverWait(browser, 5).until(element_present)
    except TimeoutException:
        print("Timed out waiting for page to load")
    finally:
        password = browser.find_element_by_id("loginPassword")
        button = browser.find_element_by_id("loginSubmit")
        email.send_keys(username)
        password.send_keys(user_password)
        time.sleep(2)
        button.click()
        time.sleep(2)
        browser.get(giveaway_url)
        time.sleep(2)
        try:
            element_present = EC.presence_of_element_located((By.XPATH, '//i[@class="fas fa-calendar-alt daily"]'))
            daily_click = WebDriverWait(browser, 10).until(element_present)
            daily_click.click()
            time.sleep(2)
            daily_click.click()
            time.sleep(2)
            browser.quit()
        except TimeoutException:
            print("Timed out waiting for page to load")

def daily_mission():
    selhdrs=["email","password"]
    user_list =  excel_to_dict('data.xlsx',selhdrs)

    airtable = Airtable('appOHsYryKh7fcZsF', 'urlTable',api_key="keyrErNCsKQjt5BB2")
    airtable_dict = airtable.get_all(view='elamUrls', sort='url')
    url_list = []
    for row in airtable_dict:
        if row["fields"]:
            url_list.append(row["fields"]["url"])

    for user in user_list:
        for url in url_list:
            go_to_givelab(user["email"],user["password"],url)
            time.sleep(10)

schedule.every(1450).minutes.do(daily_mission)


while True:
    schedule.run_pending()
    time.sleep(30)
